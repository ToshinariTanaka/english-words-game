#!/usr/bin/env python3
"""Fill blank study-app Excel choices with OpenAI-generated distractors.

Reads an .xlsx file with columns:
row_number,level,question,correct,choice1,choice2,choice3,total_correct,total_wrong,accuracy,current_streak,note

Rows whose choice1-choice3 cells are all blank are processed in batches. Each batch is
sent to the OpenAI API, and valid distractors are written back by row_number.
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sys
import time
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

REQUIRED_COLUMNS = [
    "row_number",
    "level",
    "question",
    "correct",
    "choice1",
    "choice2",
    "choice3",
    "total_correct",
    "total_wrong",
    "accuracy",
    "current_streak",
    "note",
]
CHOICE_COLUMNS = ["choice1", "choice2", "choice3"]
DEFAULT_MODEL = "gpt-4.1-mini"


@dataclass
class RowItem:
    excel_row: int
    row_number: str
    level: str
    question: str
    correct: str


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalize_for_duplicate(value: str) -> str:
    return " ".join(value.strip().casefold().split())


def contains_japanese(value: str) -> bool:
    return any(
        "\u3040" <= char <= "\u30ff" or "\u3400" <= char <= "\u9fff"
        for char in value
    )


def contains_ascii_letter(value: str) -> bool:
    return any("a" <= char.lower() <= "z" for char in value)


def is_blank(value: object) -> bool:
    return clean(value) == ""


def find_header(ws) -> dict[str, int]:
    headers = {clean(cell.value): index for index, cell in enumerate(ws[1], start=1)}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError(f"必須列が見つかりません: {', '.join(missing)}")
    return headers


def collect_unfilled_rows(ws, headers: dict[str, int]) -> list[RowItem]:
    rows: list[RowItem] = []
    for excel_row in range(2, ws.max_row + 1):
        if all(is_blank(ws.cell(excel_row, headers[column]).value) for column in CHOICE_COLUMNS):
            row_number = clean(ws.cell(excel_row, headers["row_number"]).value)
            if not row_number:
                continue
            rows.append(
                RowItem(
                    excel_row=excel_row,
                    row_number=row_number,
                    level=clean(ws.cell(excel_row, headers["level"]).value),
                    question=clean(ws.cell(excel_row, headers["question"]).value),
                    correct=clean(ws.cell(excel_row, headers["correct"]).value),
                )
            )
    return rows


def chunked(items: list[RowItem], size: int) -> Iterable[list[RowItem]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]


def build_prompt(rows: list[RowItem]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(["row_number", "level", "question", "correct"])
    for row in rows:
        writer.writerow([row.row_number, row.level, row.question, row.correct])

    return f"""あなたは英語学習アプリ用の4択問題データを補完する教材編集者です。
以下のCSVの各行について、correct と重複しない不正解選択肢を3つ作ってください。

要件:
- 出力はCSVのみ。説明文、Markdown、コードフェンスは禁止。
- ヘッダーは必ず row_number,choice1,choice2,choice3 にしてください。
- row_number は入力と完全一致させてください。
- choice1〜choice3 は空欄禁止です。
- choice1〜choice3 は必ず correct と同じ言語で作ってください。
- correct が日本語を含む場合、choice1〜choice3 は必ず日本語のみで出力してください。
- correct が日本語を含む場合、choice1〜choice3 に英単語、英語フレーズ、英語類義語を入れることは禁止です。
- correct が日本語を含む場合、choice1〜choice3 に A-Z または a-z が1文字でも含まれていたら不正です。
- question が英単語で correct が日本語の場合は、日本語4択問題として扱い、選択肢は日本語の不正解訳だけにしてください。
- 悪い例:
  question: ad
  correct: 広告
  choice1: advertisement
  choice2: notice
  choice3: announcement
- 良い例:
  question: ad
  correct: 広告
  choice1: 住所
  choice2: 案内
  choice3: 発表
- correct と choice1〜choice3 は重複禁止です。
- choice1〜choice3 同士も重複禁止です。
- correct の単なる類義語、言い換え、ほぼ同義の近縁語は禁止です。正解として成立しそうな語は選ばないでください。
- correct の派生語、複合語、接頭辞・接尾辞を付けただけの語は禁止です。
- correct の文字列を含む語は禁止です。例: correct が「中毒」の場合、「中毒症」「アルコール中毒」は禁止です。
- 「中毒→依存→依存症」のように、同じ意味領域の近縁語を連発しないでください。
- choice1〜choice3 は意味領域を分散させてください。3つとも感情、3つとも病気、3つとも行動などに偏らせないでください。
- 学習者が混同しやすいが、別概念で明確に不正解の語を優先してください。品詞・抽象度・長さは近くても、意味は正解からずらしてください。
- 悪い例:
  correct: 中毒
  choice1: 依存
  choice2: 依存症
  choice3: 中毒症
- 良い例:
  correct: 中毒
  choice1: 習慣
  choice2: 興味
  choice3: 熱意
- カンマや改行が必要な値はRFC 4180形式でCSVクォートしてください。

入力CSV:
{buffer.getvalue()}"""


def call_openai(prompt: str, model: str, api_key: str, timeout: int) -> str:
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": "Return only valid CSV."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.7,
    }
    request = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"OpenAI API error HTTP {exc.code}: {body}") from exc
    return data["choices"][0]["message"]["content"]


def strip_code_fence(text: str) -> str:
    stripped = text.strip()
    if stripped.startswith("```"):
        lines = stripped.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        return "\n".join(lines).strip()
    return stripped


def parse_response(text: str) -> dict[str, dict[str, str]]:
    reader = csv.DictReader(io.StringIO(strip_code_fence(text)))
    if reader.fieldnames != ["row_number", "choice1", "choice2", "choice3"]:
        raise ValueError(f"AI出力CSVのヘッダーが不正です: {reader.fieldnames}")
    parsed: dict[str, dict[str, str]] = {}
    for line_number, row in enumerate(reader, start=2):
        row_number = clean(row.get("row_number"))
        if not row_number:
            raise ValueError(f"AI出力CSV {line_number}行目のrow_numberが空です")
        parsed[row_number] = {column: clean(row.get(column)) for column in CHOICE_COLUMNS}
    return parsed


def validate_choices(item: RowItem, choices: dict[str, str]) -> list[str]:
    errors: list[str] = []
    values = [choices.get(column, "") for column in CHOICE_COLUMNS]
    if any(not value for value in values):
        errors.append("choice1〜choice3に空欄があります")
    normalized_correct = normalize_for_duplicate(item.correct)
    normalized_choices = [normalize_for_duplicate(value) for value in values]
    if normalized_correct and normalized_correct in normalized_choices:
        errors.append("correct と choice1〜choice3 が重複しています")
    if contains_japanese(item.correct) and any(contains_ascii_letter(value) for value in values):
        errors.append("correct が日本語を含むため choice1〜choice3 に英字[A-Za-z]を含めないでください")
    if len([value for value in normalized_choices if value]) != len(set(value for value in normalized_choices if value)):
        errors.append("choice1〜choice3 同士が重複しています")
    return errors


def validate_batch(rows: list[RowItem], parsed: dict[str, dict[str, str]]) -> dict[str, list[str]]:
    errors: dict[str, list[str]] = {}
    expected = {row.row_number for row in rows}
    missing = expected - set(parsed)
    for row_number in missing:
        errors.setdefault(row_number, []).append("AI出力にrow_numberがありません")
    for row in rows:
        if row.row_number in parsed:
            row_errors = validate_choices(row, parsed[row.row_number])
            if row_errors:
                errors[row.row_number] = row_errors
    return errors


def apply_choices(ws, headers: dict[str, int], rows: list[RowItem], parsed: dict[str, dict[str, str]]) -> None:
    by_row_number = {row.row_number: row for row in rows}
    for row_number, choices in parsed.items():
        item = by_row_number.get(row_number)
        if item is None:
            continue
        for column in CHOICE_COLUMNS:
            ws.cell(item.excel_row, headers[column]).value = choices[column]


def process_workbook(args: argparse.Namespace) -> int:
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key and not args.dry_run:
        raise RuntimeError("環境変数 OPENAI_API_KEY が設定されていません")

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised in user environment
        raise RuntimeError(
            "openpyxl が見つかりません。先に `python3 -m pip install -r tools/requirements.txt` を実行してください。"
        ) from exc

    workbook = load_workbook(args.input)
    ws = workbook[args.sheet] if args.sheet else workbook.active
    headers = find_header(ws)
    unfilled = collect_unfilled_rows(ws, headers)
    output = Path(args.output)
    print(f"未補完行: {len(unfilled)} 件")

    if args.dry_run:
        print("dry-run のためAPI呼び出しと保存は行いません")
        return 0

    processed = 0
    try:
        for batch_number, batch in enumerate(chunked(unfilled, args.batch_size), start=1):
            pending = batch
            last_errors: dict[str, list[str]] = {}
            for attempt in range(1, args.max_retries + 1):
                print(f"Batch {batch_number}: {len(pending)}件を処理中 (attempt {attempt}/{args.max_retries})")
                response = call_openai(build_prompt(pending), args.model, api_key, args.timeout)
                parsed = parse_response(response)
                errors = validate_batch(pending, parsed)
                valid_rows = [row for row in pending if row.row_number in parsed and row.row_number not in errors]
                apply_choices(ws, headers, valid_rows, parsed)
                last_errors = errors
                pending = [row for row in pending if row.row_number in errors]
                if not pending:
                    break
                print(f"  再試行対象: {len(pending)}件")
                time.sleep(args.retry_delay)
            if pending:
                row_by_number = {row.row_number: row for row in batch}
                details = []
                for row_number, errs in last_errors.items():
                    row = row_by_number.get(row_number)
                    choices = parsed.get(row_number, {}) if "parsed" in locals() else {}
                    row_label = f"Excel行{row.excel_row}" if row else "Excel行不明"
                    content = (
                        f"question={row.question!r}, correct={row.correct!r}, "
                        f"choice1={choices.get('choice1', '')!r}, "
                        f"choice2={choices.get('choice2', '')!r}, "
                        f"choice3={choices.get('choice3', '')!r}"
                        if row
                        else f"choices={choices!r}"
                    )
                    details.append(f"row_number={row_number} ({row_label}): {', '.join(errs)}; {content}")
                detail = " / ".join(details)
                raise RuntimeError(f"最大再試行回数を超えても不正な選択肢が残りました: {detail}")
            processed += len(batch)
            workbook.save(output)
            print(f"途中保存: {output} ({processed}/{len(unfilled)}件完了)")
    except Exception:
        if processed > 0:
            workbook.save(output)
            print(f"エラー発生のため処理済み部分を保存しました: {output}", file=sys.stderr)
        raise

    workbook.save(output)
    print(f"完成ファイルを保存しました: {output}")
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="study-app用Excelのchoice1〜choice3をOpenAI APIで自動補完します。")
    parser.add_argument("input", help="入力 .xlsx ファイル")
    parser.add_argument("-o", "--output", default="output_completed.xlsx", help="出力 .xlsx ファイル")
    parser.add_argument("--sheet", help="対象シート名。省略時はアクティブシート")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"OpenAIモデル名 (default: {DEFAULT_MODEL})")
    parser.add_argument("--batch-size", type=int, default=50, help="1回のAPI呼び出しで処理する行数")
    parser.add_argument("--max-retries", type=int, default=3, help="重複・空欄行の最大再試行回数")
    parser.add_argument("--retry-delay", type=float, default=1.0, help="再試行前の待機秒数")
    parser.add_argument("--timeout", type=int, default=120, help="APIタイムアウト秒数")
    parser.add_argument("--dry-run", action="store_true", help="未補完行数と列だけ確認し、API呼び出し・保存をしない")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    if args.batch_size <= 0:
        raise SystemExit("--batch-size は1以上にしてください")
    if args.max_retries <= 0:
        raise SystemExit("--max-retries は1以上にしてください")
    return process_workbook(args)


if __name__ == "__main__":
    raise SystemExit(main())
