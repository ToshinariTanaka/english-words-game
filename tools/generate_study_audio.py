#!/usr/bin/env python3
"""Generate study-app MP3 audio files from the official 4-sheet Excel workbook.

The tool reads column C (question) and column M (question_key) from each official
study-app worksheet, validates rows that are complete enough to be quiz items,
and writes {question_key}.mp3 files into a local output directory.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - exercised only in missing dependency envs
    raise SystemExit(
        "openpyxl が見つかりません。先に `python3 -m pip install -r tools/requirements.txt` を実行してください。"
    ) from exc

SHEET_CONFIGS = {
    "word": {"sheet": "★英単語", "prefix": "w", "pattern": re.compile(r"^w\d{6}$")},
    "chunk": {"sheet": "★チャンク", "prefix": "c", "pattern": re.compile(r"^c\d{6}$")},
    "phrase": {"sheet": "★文節和訳", "prefix": "p", "pattern": re.compile(r"^p\d{6}$")},
    "definition": {"sheet": "★英文和訳", "prefix": "s", "pattern": re.compile(r"^s\d{6}$")},
}
MODE_CHOICES = ["all", *SHEET_CONFIGS.keys()]
DEFAULT_OUTPUT_DIR = "audio_output"
LOG_FILENAME = "generation_log.csv"
OPENAI_TTS_ENDPOINT = "https://api.openai.com/v1/audio/speech"
DEFAULT_OPENAI_TTS_MODEL = "gpt-4o-mini-tts"
DEFAULT_OPENAI_TTS_VOICE = "alloy"


@dataclass(frozen=True)
class AudioItem:
    mode: str
    sheet_name: str
    excel_row: int
    question: str
    question_key: str


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="study-app正式4シートExcelから question_key.mp3 を一括生成します。"
    )
    parser.add_argument("input", type=Path, help="入力するstudy-app正式形式の .xlsx ファイル")
    parser.add_argument("--mode", choices=MODE_CHOICES, default="all", help="生成対象モード（既定: all）")
    parser.add_argument("--output", type=Path, default=Path(DEFAULT_OUTPUT_DIR), help="出力フォルダ（既定: audio_output）")
    parser.add_argument("--overwrite", action="store_true", help="既存MP3を上書き生成します")
    parser.add_argument("--limit", type=int, help="最大生成件数を指定します")
    parser.add_argument("--start-key", help="この question_key 以降だけを対象にします（例: s000001）")
    parser.add_argument("--end-key", help="この question_key 以前だけを対象にします（例: s000050）")
    parser.add_argument("--dry-run", action="store_true", help="MP3を作らず生成対象一覧だけをログに出します")
    parser.add_argument("--timeout", type=int, default=60, help="TTS API通信タイムアウト秒（既定: 60）")
    return parser.parse_args()


def selected_modes(mode: str) -> list[str]:
    return list(SHEET_CONFIGS) if mode == "all" else [mode]


def validate_range_key(key: str, allowed_modes: list[str], label: str) -> None:
    if not key:
        return
    if not any(SHEET_CONFIGS[mode]["pattern"].match(key) for mode in allowed_modes):
        allowed = ", ".join(f"{SHEET_CONFIGS[mode]['prefix']}000001" for mode in allowed_modes)
        raise ValueError(f"{label} は選択モードの形式に一致しません: {key}（例: {allowed}）")


def has_quiz_choices(ws, row: int) -> bool:
    # D〜G列: correct, choice1, choice2, choice3
    return all(clean(ws.cell(row=row, column=column).value) for column in range(4, 8))


def collect_items(workbook_path: Path, modes: list[str]) -> list[AudioItem]:
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    missing_sheets = [SHEET_CONFIGS[mode]["sheet"] for mode in modes if SHEET_CONFIGS[mode]["sheet"] not in wb.sheetnames]
    if missing_sheets:
        raise ValueError(f"必須シートが見つかりません: {', '.join(missing_sheets)}")

    items: list[AudioItem] = []
    for mode in modes:
        config = SHEET_CONFIGS[mode]
        ws = wb[config["sheet"]]
        for row in range(2, ws.max_row + 1):
            question = clean(ws.cell(row=row, column=3).value)
            question_key = clean(ws.cell(row=row, column=13).value)
            if not question or not question_key or not has_quiz_choices(ws, row):
                continue
            if not config["pattern"].match(question_key):
                continue
            items.append(
                AudioItem(
                    mode=mode,
                    sheet_name=config["sheet"],
                    excel_row=row,
                    question=question,
                    question_key=question_key,
                )
            )
    return items


def in_key_range(item: AudioItem, start_key: str | None, end_key: str | None) -> bool:
    if start_key and item.question_key < start_key:
        return False
    if end_key and item.question_key > end_key:
        return False
    return True


def synthesize_text_to_mp3(text: str, output_path: Path, timeout: int = 60) -> None:
    """Provider boundary for TTS generation.

    The current implementation uses OpenAI's speech endpoint when OPENAI_API_KEY is
    available. Keep API keys in environment variables only; never put them in browser
    code, committed files, or command-line arguments.
    """

    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY が未設定です。--dry-run を使うか環境変数を設定してください。")

    payload = {
        "model": os.environ.get("OPENAI_TTS_MODEL", DEFAULT_OPENAI_TTS_MODEL),
        "voice": os.environ.get("OPENAI_TTS_VOICE", DEFAULT_OPENAI_TTS_VOICE),
        "input": text,
        "response_format": "mp3",
    }
    request = urllib.request.Request(
        OPENAI_TTS_ENDPOINT,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            output_path.write_bytes(response.read())
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"TTS API error HTTP {exc.code}: {body}") from exc


def write_log(log_path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = ["question_key", "mode", "question", "output_file", "status", "message"]
    with log_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    args = parse_args()
    if args.limit is not None and args.limit < 1:
        print("--limit は1以上を指定してください。", file=sys.stderr)
        return 2

    modes = selected_modes(args.mode)
    try:
        validate_range_key(args.start_key, modes, "--start-key")
        validate_range_key(args.end_key, modes, "--end-key")
        if args.start_key and args.end_key and args.start_key > args.end_key:
            raise ValueError("--start-key は --end-key 以下にしてください。")
        items = collect_items(args.input, modes)
    except Exception as exc:
        print(f"エラー: {exc}", file=sys.stderr)
        return 1

    filtered = [item for item in items if in_key_range(item, args.start_key, args.end_key)]
    if args.limit is not None:
        filtered = filtered[: args.limit]

    args.output.mkdir(parents=True, exist_ok=True)
    log_rows: list[dict[str, str]] = []
    generated = skipped = failed = dry_run = 0

    for item in filtered:
        output_file = args.output / f"{item.question_key}.mp3"
        row = {
            "question_key": item.question_key,
            "mode": item.mode,
            "question": item.question,
            "output_file": str(output_file),
            "status": "",
            "message": "",
        }
        if args.dry_run:
            row["status"] = "dry-run"
            row["message"] = "生成対象です。--dry-run のためMP3は作成していません。"
            dry_run += 1
        elif output_file.exists() and not args.overwrite:
            row["status"] = "skipped"
            row["message"] = "同名MP3が既に存在します。上書きする場合は --overwrite を指定してください。"
            skipped += 1
        else:
            try:
                synthesize_text_to_mp3(item.question, output_file, timeout=args.timeout)
                row["status"] = "generated"
                row["message"] = "生成しました。"
                generated += 1
            except Exception as exc:  # keep processing other rows and record failures
                row["status"] = "failed"
                row["message"] = str(exc)
                failed += 1
        log_rows.append(row)

    log_path = args.output / LOG_FILENAME
    write_log(log_path, log_rows)
    print(
        f"対象 {len(filtered)} 件 / generated={generated} skipped={skipped} dry-run={dry_run} failed={failed} / log={log_path}"
    )
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
